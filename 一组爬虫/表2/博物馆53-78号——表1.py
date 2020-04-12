#-*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配
import urllib.request,urllib.error # 制定URL，后去网页数据
import xlwt  # 进行excel操作
import sqlite3  # 进行SQLite数据库操作
import openpyxl
from openpyxl import load_workbook
import xlrd

# 创建正则表达式对象，表示规则（字符串的模式）
# 部分通用的正则表达式
tongyong = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 博物馆名称  museumName
museumName = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 博物馆地点  museumPosition
museumPosition = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 博物馆类别  buildingType
buildingType = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 展区数量  number
number = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 建馆时间  buildingTime
buildingTime = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)

# 开放时间  openingHours
openingHours = re.compile(r'<dd class="basicInfo-item value">(.*?)</dd>', re.S)
# 图片  picture
picture = re.compile(r'<img src="(.*?)"', re.S)


def getexcel():
    # 打开文件
    web = str(r'D:\python_code\\博物馆链接总表 .xlsx')
    sheet = xlrd.open_workbook(web)
    table = sheet.sheet_by_name('Sheet1')
    # 查询工作表
    sourse = []  # 存储从表里拿出来的数据
    # 第一列的 0-25 个数据
    for i in range(52,78):
        # t =[]
        # # print(table.cell(i,0).value)
        # #t.append('"')
        # t.append(table.cell(i,0).value) # i——行，0——列
        # #t.append('"')
        # sourse.append("".join(t))
        sourse.append(table.cell(i,0).value) # i——行，0——列
        # print(table.cell(i,0).value)
    return sourse

def main():
    # 获取execl表的数据
    sourse = getexcel()
    # baseurl = "https://movie.douban.com/top250?start="

    # 创建workbook对象
    book = xlwt.Workbook(encoding="utf-8", style_compression=0)
    # 创建工作表
    sheet1 = book.add_sheet('博物馆基础信息表', cell_overwrite_ok=True)
    sheet4 = book.add_sheet('博物馆展览表', cell_overwrite_ok=True)
    sheet5 = book.add_sheet('博物馆经典藏品表', cell_overwrite_ok=True)


            # 博物馆基础信息表
    # col = ("博物馆编号", "博物馆名称", "博物馆地点", "博物馆类别", "展区数量",
    #        # 简介表
    #        "博物馆编号", "建馆时间", "发展历史", "博物馆图片", "宣传视频", "宣传音频",
    #        # 参观信息表
    #        "博物馆编号", "门票", "开放时间", "建筑布局", "藏品数量", "藏品种类",
    #        # 展览表
    #        "博物馆编号", "展览编号", "主题", "时间", "介绍", "地点",
    #        # 经典藏品表
    #        "博物馆编号", "藏品编号", "名称", "名称","经典藏品图片", "年代", "说明", "尺寸", "来源", "质地", "出土时间", "文物等级",

           # 博物馆基础信息表
    # col = ("博物馆编号", "博物馆名称", "博物馆地点", "博物馆类别", "展区数量", "建馆时间", "图片", "宣传视频或音频", "门票", "开放时间", "藏品数量", "藏品种类")
    col = ("博物馆编号", "博物馆名称", "博物馆地点", "博物馆类别", "展区数量","建馆时间", "图片", "用户视频或音频", "门票", "开放时间", "藏品数量", "藏品种类","评价", "评分")
    print(col)
    for i in range(0, 14):
        sheet1.write(0, i, col[i])  # 列名

    # 第53个博物馆——宁波博物馆
    baseurl = sourse[0]
    datalist = getData53(baseurl,53)
    print("datalist = ",datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath,book,sheet1,0)
    # askURL(sourse[0])

    # 第54个博物馆——杭州博物馆
    baseurl = sourse[1]
    # 爬取博物馆基础信息表
    datalist = getData54(baseurl, 54)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 1)
    # askURL(sourse[1])

    # 第55个博物馆——温州博物馆
    baseurl = sourse[2]
    # 爬取博物馆基础信息表
    datalist = getData55(baseurl, 55)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 2)
    # askURL(sourse[2])

    # 第56个博物馆——安徽博物院
    baseurl = sourse[3]
    # 爬取博物馆基础信息表
    datalist = getData56(baseurl, 56)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 3)
    # askURL(sourse[3])

    # 第57个博物馆——安徽中国徽州文化博物馆
    baseurl = sourse[4]
    # 爬取博物馆基础信息表
    datalist = getData57(baseurl, 57)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 4)
    # askURL(sourse[4])

    # 第58个博物馆——福建博物院
    baseurl = sourse[5]
    # 爬取博物馆基础信息表
    datalist = getData58(baseurl, 58)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 5)
    # askURL(sourse[5])

    # 第59个博物馆——古田会议纪念馆
    baseurl = sourse[6]
    # 爬取博物馆基础信息表
    datalist = getData59(baseurl, 59)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 6)
    # askURL(sourse[6])


    # 第60个博物馆——泉州海外交通史博物馆
    baseurl = sourse[7]
    # 爬取博物馆基础信息表
    datalist = getData60(baseurl, 60)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 7)
    # askURL(sourse[7])

    # 第61个博物馆——中国闽台缘博物馆
    baseurl = sourse[8]
    # 爬取博物馆基础信息表
    datalist = getData61(baseurl, 61)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 8)
    # askURL(sourse[8])

    # 第62个博物馆——中央苏区（闽西）历史博物馆
    baseurl = sourse[9]
    # 爬取博物馆基础信息表
    datalist = getData62(baseurl, 62)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 9)
    # askURL(sourse[9])

    # 第63个博物馆——井冈山革命博物馆
    baseurl = sourse[10]
    # 爬取博物馆基础信息表
    datalist = getData63(baseurl, 63)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 10)
    # askURL(sourse[10])

    # 第64个博物馆——江西省博物馆
    baseurl = sourse[11]
    # 爬取博物馆基础信息表
    datalist = getData64(baseurl, 64)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 11)
    # askURL(sourse[11])

    # 第65个博物馆——瑞金中央革命根据地纪念馆
    baseurl = sourse[12]
    # 爬取博物馆基础信息表
    datalist = getData65(baseurl, 65)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 12)
    # askURL(sourse[12])
    # 第66个博物馆——南昌八一起义纪念馆
    baseurl = sourse[13]
    # 爬取博物馆基础信息表
    datalist = getData66(baseurl, 66)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 13)
    # askURL(sourse[13])
    # 第67个博物馆——安源路矿工人运动纪念馆
    baseurl = sourse[14]
    # 爬取博物馆基础信息表
    datalist = getData67(baseurl, 67)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 14)
    # askURL(sourse[14])
    # 第68个博物馆——青岛市博物馆
    baseurl = sourse[15]
    # 爬取博物馆基础信息表
    datalist = getData68(baseurl, 68)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 15)
    # askURL(sourse[15])
    # 第69个博物馆——中国甲午战争博物馆
    baseurl = sourse[16]
    # 爬取博物馆基础信息表
    datalist = getData69(baseurl, 69)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 16)
    # askURL(sourse[16])
    # 第70个博物馆——青州博物馆
    baseurl = sourse[17]
    # 爬取博物馆基础信息表
    datalist = getData70(baseurl, 70)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 17)
    # askURL(sourse[17])
    # 第71个博物馆——山东博物馆
    baseurl = sourse[18]
    # 爬取博物馆基础信息表
    datalist = getData71(baseurl, 71)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 18)
    # askURL(sourse[18])
    # 第72个博物馆——烟台市博物馆
    baseurl = sourse[19]
    # 爬取博物馆基础信息表
    datalist = getData72(baseurl, 72)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 19)
    # askURL(sourse[19])
    # 第73个博物馆——潍坊市博物馆
    baseurl = sourse[20]
    # 爬取博物馆基础信息表
    datalist = getData73(baseurl, 73)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 20)
    # askURL(sourse[20])
    # 第74个博物馆——河南博物院
    baseurl = sourse[21]
    # 爬取博物馆基础信息表
    datalist = getData74(baseurl, 74)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 21)
    # askURL(sourse[21])
    # 第75个博物馆——郑州博物馆 （嵩山南路馆区）
    baseurl = sourse[22]
    # 爬取博物馆基础信息表
    datalist = getData75(baseurl, 75)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 22)
    # askURL(sourse[22])
    # 第76个博物馆——洛阳博物馆
    baseurl = sourse[23]
    # 爬取博物馆基础信息表
    datalist = getData76(baseurl, 76)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 23)
    # askURL(sourse[23])
    # 第77个博物馆——南阳市汉画馆
    baseurl = sourse[24]
    # 爬取博物馆基础信息表
    datalist = getData77(baseurl, 77)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 24)
    # askURL(sourse[24])
    # 第78个博物馆——开封市博物馆
    baseurl = sourse[25]
    # 爬取博物馆基础信息表
    datalist = getData78(baseurl, 78)
    print("datalist = ", datalist)
    savepath = '博物馆信息.xlsx'  # 存在excel里
    saveData(datalist, savepath, book, sheet1, 25)
    # askURL(sourse[25])

# 爬取网页
def getData53(baseurl,i):

    # 宣传视频或音频  VideoAudio

    # 门票  ticket
    ticket = re.compile(r'</div></a>(.*?)</div>', re.S)
    # 藏品数量  collectionNum
    collectionNum = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
    # 藏品种类  collectionType
    collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)

    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0,'0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del(name[0])
        lenname = len(name)
        del(name[lenname-1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("3个")

        # 建馆时间  buildingTime
        Time = re.findall(buildingType, item)[4]
        Time = list(Time)
        del (Time[0])
        lenname = len(Time)
        del (Time[lenname - 1])
        datalist.append("".join(Time))
        # datalist.append("Null")

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)



        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
        # 门票  ticket
    datalist.append("每日发放免费参观券3000张（含预约观众500人），发完为止。")
    # for item in soup.find_all('div', class_="para"):
    #     item = str(item)
    #     tic = re.findall(ticket, item)[0]
    #     datalist.append(tic)

        # 开放时间  openingHours
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        Hours = re.findall(openingHours, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])
        datalist.append("".join(Hours))

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        # 藏品数量  collectionNum
        collNum = re.findall(collectionNum, item)[1]
        conum = collNum[69:73]
        datalist.append(conum)

        # 藏品种类  collectionType
        collType = re.findall(collectionType, item)[1]
        cotype = collNum[42:66]
        datalist.append(cotype)

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData54(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0,'0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del(name[0])
        lenname = len(name)
        del(name[lenname-1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 门票  ticket
        ticket = re.findall(buildingType, item)[6]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])



        # 展区数量
        datalist.append("6个")

        # 建馆时间  buildingTime
    for item in soup.find_all('div', class_="main-content"):
        item = str(item)
        buildingTime = re.compile(r'<div class="para" label-module="para">(.*?)</div>')
        # 藏品数量  collectionNum
        bTime = re.findall(buildingTime, item)[1]
        Time = bTime[6:11]
        datalist.append(Time)

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)



        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
        # 门票  ticket
    datalist.append(ticket)
    # for item in soup.find_all('div', class_="para"):
    #     item = str(item)
    #     tic = re.findall(ticket, item)[0]
    #     datalist.append(tic)

        # 开放时间  openingHours
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        Hours = re.findall(tongyong, item)[4]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])
        datalist.append("".join(Hours))

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[0]
        # 藏品数量  collectionNum
        collNum = tt[385:388]
        datalist.append(collNum)

        # 藏品种类  collectionType
        cotype = tt[392:409]
        datalist.append(cotype)

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData55(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("七")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))


    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    # 门票  ticket
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        ticket = re.findall(tongyong, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])
        datalist.append("".join(ticket))
    # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])
        datalist.append("".join(Hours))

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[1]

        # 藏品数量  collectionNum
        collNum = tt[203:208]
        datalist.append(collNum)
        # print(collNum)

        # 藏品种类  collectionType
        cotype = tt[211:231]
        datalist.append(cotype)
        # print(cotype)
        # print(tt)


    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData56(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("七")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")

    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # 门票  ticket
        datalist.append("免费")

    # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])
        datalist.append("".join(Hours))

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[0]

        # 展区数量 -- 4
        num = tt[292:295]
        datalist[4] = num

        # 藏品数量  collectionNum
        collNum = tt[300:307]
        datalist.append(collNum)


    for item in soup.find_all('div', class_="main-content"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[28]
        t = tt[41:64]
        # 藏品种类  collectionType
        string = []
        string.append(t)
        t = tt[133:137] + '、'
        string.append(t)
        t = tt[211:215] + '、'
        string.append(t)
        t = tt[220:236]
        string.append(t)
        datalist.append("".join(string))

        # 门票  ticket
        tt = re.findall(collectionType, item)[74]
        t = tt[5:7]
        datalist[8] = t

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")


    # print("  datalist = ",datalist)
    return datalist

def getData57(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[2]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[1]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("6个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[3]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[4]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])

        # 藏品数量  collectionNum
        collectionNum = re.findall(tongyong, item)[9]
        collectionNum = list(collectionNum)
        del (collectionNum[0])
        lenname = len(collectionNum)
        del (collectionNum[lenname - 1])

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    datalist.append("".join(ticket))
    datalist.append("".join(Hours))
    datalist.append("".join(collectionNum))



    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[1]
        # 藏品种类  collectionType
        collectionType = tt[22:58]
        datalist.append(collectionType)
    # print("  datalist = ",datalist)

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    return datalist

def getData58(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        tt = re.findall(museumPosition, item)[3]
        t = tt[1:7]
        Position = []
        Position.append(t)
        t = tt[67:70]
        Position.append(t)
        t = tt[165:168]
        Position.append(t)
        t = tt[172:175]
        Position.append(t)
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("Null")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[7]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    datalist.append("".join(ticket))
    datalist.append("".join(Hours))





    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[0]
        # 展区数量
        t = tt[241:244]
        datalist[4] = t

        # 藏品数量  collectionNum
        tt = re.findall(collectionType, item)[1]
        t = tt[25:31]
        datalist.append(t)

        # 藏品种类  collectionType
        datalist.append("陶器、禽类骨骼、石器、石片、骨器、瓷器、书画")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData59(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[1]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        datalist.append("国家一级博物馆 ")

        # 展区数量
        datalist.append("Null")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[3]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[4]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("Null")

    # 藏品数量  collectionNum
    for item in soup.find_all('div', class_="main-content"):
        item = str(item)
        xx = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(xx, item)[8]
        # 博物馆类别
        t = tt[6:10]
        datalist.append(t)

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        collectionType = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(collectionType, item)[0]
        # 博物馆类别
        t = tt[158:165]
        datalist[3] = t

        # 藏品种类  collectionType
        datalist.append("旧址、书籍")

        # 评价  evaluate
        datalist.append("环境优美，馆藏丰富。")
        # 评分  grade
        datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData60(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("11个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])


    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    datalist.append("".join(ticket))
    datalist.append("".join(Hours))

    # 藏品数量  collectionNum
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("锚具、石刻、陶瓷器、船模、器物")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData61(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("3个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    datalist.append("".join(ticket))
    datalist.append("".join(Hours))

    # 藏品数量  collectionNum
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("石壁雕、火药爆绘壁画、陶器、编钟")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData62(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[2]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[1]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("9个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[3]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

    for item in soup.find_all('div', class_="lemma-summary"):
        item = str(item)
        xx = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(xx, item)[1]
        # 展区数量
        t = tt[54:57]
        datalist[4] = t

    # 宣传视频或音频  VideoAudio
    datalist.append("Null")

    for item in soup.find_all('div', class_="main-content"):
        item = str(item)
        xx = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        # 门票  ticket
        tt = re.findall(xx, item)[54]
        t = tt[2:4]
        datalist.append(t)

        # 开放时间  openingHours
        tt = re.findall(xx, item)[51]
        string = []
        string.append(tt)
        tt = re.findall(xx, item)[52]
        string.append(tt)
        tt = re.findall(xx, item)[53]
        t = tt[0:34]
        string.append(t)
        datalist.append("".join(string))

    # 藏品数量  collectionNum
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("书籍、服饰、瓷器、物件")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData63(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("19个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[4]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])
        datalist.append("".join(bTime))

        # 门票  ticket
        ticket = re.findall(tongyong, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[5]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    datalist.append("".join(ticket))
    datalist.append("".join(Hours))

    # 藏品数量  collectionNum
    datalist.append("3万余件")
    # 藏品种类  collectionType
    datalist.append("墨宝珍迹、影视资料、书籍、手稿、武器")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData64(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = re.findall(museumPosition, item)[3]
        Position = list(Position)
        del (Position[0])
        lenname = len(Position)
        del (Position[lenname - 1])
        datalist.append("".join(Position))

        # 博物馆类别
        Type = re.findall(buildingType, item)[2]
        Type = list(Type)
        del (Type[0])
        lenname = len(Type)
        del (Type[lenname - 1])
        datalist.append("".join(Type))

        # 展区数量
        datalist.append("10个")

        # 建馆时间  buildingTime
        bTime = re.findall(tongyong, item)[8]
        bTime = list(bTime)
        del (bTime[0])
        lenname = len(bTime)
        del (bTime[lenname - 1])

        t = re.findall(tongyong, item)[9]
        t = list(t)
        del (t[0])
        lenname = len(t)
        del (t[lenname - 1])
        tt =[]
        tt.append("".join(bTime))
        tt.append("".join(t))
        tt.append(" (新馆开馆)")
        datalist.append("".join(tt))

        # 开放时间  openingHours
        Hours = re.findall(tongyong, item)[4]
        Hours = list(Hours)
        del (Hours[0])
        lenname = len(Hours)
        del (Hours[lenname - 1])

    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    ticket = "持有效证件换取门票，免费参观"
    datalist.append(ticket)
    datalist.append("".join(Hours))

    # 藏品数量  collectionNum
    datalist.append("3.4万余件")
    # 藏品种类  collectionType
    datalist.append("青铜器、金银器、陶瓷器、玉石、字画、生物标本")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData65(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点
        Position = '江西省瑞金市区的象湖镇'
        datalist.append(Position)

        # 藏品数量  collectionNum
        collectionNum = re.findall(tongyong, item)[5]
        collectionNum = list(collectionNum)
        del (collectionNum[0])
        lenname = len(collectionNum)
        del (collectionNum[lenname - 1])

        # 博物馆类别
        Type = '纪念性博物馆'
        datalist.append(Type)

        # 展区数量
        datalist.append("Null")

        # 建馆时间  buildingTime
        datalist.append('1958年')




    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    ticket = "免费"
    datalist.append(ticket)

    # 开放时间  openingHours
    for item in soup.find_all('div', class_="main-content"):
        item = str(item)
        xx = re.compile(r'<div class="para" label-module="para">(.*?)</div>', re.S)
        tt = re.findall(xx, item)[51]
        # 博物馆类别
        t = []
        t.append(tt)
        tt = re.findall(xx, item)[52]
        t.append(tt)
        tt = re.findall(xx, item)[53]
        t.append(tt)
        datalist.append("".join(t))



    # 藏品数量  collectionNum
    datalist.append("".join("".join(collectionNum)))
    # 藏品种类  collectionType
    datalist.append("遗址、图书、杂志、器件、勋章")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData66(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[2]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[1]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("两层")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[3]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # # 开放时间  openingHours
        # openingHours = re.findall(museumName, item)[0]
        # openingHours = list(openingHours)
        # del (openingHours[0])
        # lenname = len(openingHours)
        # del (openingHours[lenname - 1])



    # for item in soup.find_all('div', class_="album-wrap"):
    #     item = str(item)
    #     # 图片
    #     pic = re.findall(picture, item)[0]
    #     datalist.append(pic)
    # 图片
    datalist.append('https://bkimg.cdn.bcebos.com/pic/060828381f30e924629d91cc45086e061d95f716?x-bce-process=image/resize,m_lfit,w_220,h_220,limit_1')
        # 宣传视频或音频  VideoAudio
    datalist.append("Null")

    # 门票  ticket
    datalist.append("免费（每天免费发放3500张参观券）")
    # 开放时间  openingHours
    datalist.append('9:00-17:30（一般下午四点就不许进了，周一闭馆），9:10-11:00；13:30-15:30每半小时有讲解及设备播放。')

    # 藏品数量  collectionNum
    datalist.append("972件")
    # 藏品种类  collectionType
    datalist.append("图片、图表、文物展品、艺术品等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData67(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("12个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # # 开放时间  openingHours
        # openingHours = re.findall(museumName, item)[0]
        # openingHours = list(openingHours)
        # del (openingHours[0])
        # lenname = len(openingHours)
        # del (openingHours[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    datalist.append("免费")
    # 开放时间  openingHours
    datalist.append('淡季8:00～17:00 旺季8:00～17:30 闭馆日：星期一')

    # 藏品数量  collectionNum
    datalist.append("5000余件")
    # 藏品种类  collectionType
    datalist.append("书籍、文物展品、股票凭证等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData68(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[2]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[1]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("5个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[3]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # # 开放时间  openingHours
        # openingHours = re.findall(museumName, item)[0]
        # openingHours = list(openingHours)
        # del (openingHours[0])
        # lenname = len(openingHours)
        # del (openingHours[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)
        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    datalist.append("免费（每天限定参观人数3000人）")
    # 开放时间  openingHours
    datalist.append('5月至10月：9：00—17：00（16：00停止入场）；11月至次年4月：9：00—16：30（15：30停止入场）；每周一闭馆（法定节假日除外）')


    # 藏品数量  collectionNum
    datalist.append("16万件")
    # 藏品种类  collectionType
    datalist.append("书法、绘画、陶瓷器、铜器、玉器、钱币、玺印、甲骨、竹木牙角器等三十余个门类")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")
    # print("  datalist = ",datalist)
    return datalist

def getData69(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[2]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[1]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("8个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[3]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[6]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[4]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])



    # for item in soup.find_all('div', class_="lemma-picture text-pic layout-right"):
    #     item = str(item)
    #     # 图片
    #     pic = re.findall(picture, item)[0]
    #     datalist.append(pic)
    # 图片
    datalist.append('https://bkimg.cdn.bcebos.com/pic/f703738da9773912427ba723f8198618367ae217?x-bce-process=image/resize,m_lfit,w_220,h_220,limit_1')
        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("铁锚、鱼雷、委任状、望远镜、指南针、怀表")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData70(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("10个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)
        # 宣传视频或音频  VideoAudio
    datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("四万余件")
    # 藏品种类  collectionType
    datalist.append("简史陈列、陶瓷器、玉器、青铜器、书画、古货币、碑碣、石刻、革命文物等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")
    # print("  datalist = ",datalist)
    return datalist

def getData71(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("15个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[7]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("陶瓷器、青铜器、甲骨文、陶文、封泥、玺印、简牍、汉画像石、书画、善本书等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")
    # print("  datalist = ",datalist)
    return datalist

def getData72(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("6个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # # 门票  ticket
        # ticket = re.findall(museumName, item)[7]
        # ticket = list(ticket)
        # del (ticket[0])
        # lenname = len(ticket)
        # del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])



    # for item in soup.find_all('div', class_="summary-pic"):
    #     item = str(item)
    #     # 图片
    #     pic = re.findall(picture, item)[0]
    #     datalist.append(pic)
        # 图片
        datalist.append('https://bkimg.cdn.bcebos.com/pic/29381f30e924b8997a2b05286e061d950a7bf68f?x-bce-process=image/resize,m_lfit,w_220,h_220,limit_1')
        # 宣传视频或音频  VideoAudio
    datalist.append("Null")
    # 门票  ticket
    datalist.append("免费参观")
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("5万余件")
    # 藏品种类  collectionType
    datalist.append("陶瓷器、玉石器、青铜器、铁器、书画、丝织品和杂项等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData73(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("大于4个")

        # 建馆时间  buildingTime
        # buildingTime = re.findall(museumName, item)[4]
        # buildingTime = list(buildingTime)
        # del (buildingTime[0])
        # lenname = len(buildingTime)
        # del (buildingTime[lenname - 1])
        # datalist.append("".join(buildingTime))
        datalist.append("1962年")

        # 门票  ticket
        ticket = re.findall(museumName, item)[6]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[4]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)
        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("近8万件")
    # 藏品种类  collectionType
    datalist.append("化石、陶器、瓷器、青铜、玉石、钱币、古角牙木、书画、碑拓、古籍、碑刻、造像、织绣、玺印符牌、民俗服饰及饰品、近现代革命文物等33类")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData74(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        # museumPosition = re.findall(museumName, item)[3]
        # museumPosition = list(museumPosition)
        # del (museumPosition[0])
        # lenname = len(museumPosition)
        # del (museumPosition[lenname - 1])
        # datalist.append("".join(museumPosition))
        datalist.append("河南郑州")


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("4个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[6]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[12]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[4]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])

        # 藏品数量  collectionNum
        collectionNum = re.findall(museumName, item)[7]
        collectionNum = list(collectionNum)
        del (collectionNum[0])
        lenname = len(collectionNum)
        del (collectionNum[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    # datalist.append("近8万件")
    datalist.append("".join(collectionNum))
    # 藏品种类  collectionType
    datalist.append("史前文物、商周青铜器、历代陶瓷器、玉器等")

    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData75(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))
        # datalist.append("河南郑州")


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("3个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        # ticket = re.findall(museumName, item)[12]
        # ticket = list(ticket)
        # del (ticket[0])
        # lenname = len(ticket)
        # del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])

        # 藏品数量  collectionNum
        # collectionNum = re.findall(museumName, item)[7]
        # collectionNum = list(collectionNum)
        # del (collectionNum[0])
        # lenname = len(collectionNum)
        # del (collectionNum[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    datalist.append("现场免费领取参观券，讲解收费为50元/场")
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    # datalist.append("近8万件")
    datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("编钟、青铜礼器、酒器、玉器、占卜甲骨、骨蚌器、日用陶器、旧石器时代、新石器时代文物等")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData76(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[2]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))
        # datalist.append("河南郑州")


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[14]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("8个")

        # 建馆时间  buildingTime
        # buildingTime = re.findall(museumName, item)[4]
        # buildingTime = list(buildingTime)
        # del (buildingTime[0])
        # lenname = len(buildingTime)
        # del (buildingTime[lenname - 1])
        # datalist.append("".join(buildingTime))
        datalist.append("1958年")

        # 门票  ticket
        ticket = re.findall(museumName, item)[6]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        tt = re.findall(museumName, item)[5]
        openingHours = tt[1:20]


        # 藏品数量  collectionNum
        tt = re.findall(museumName, item)[12]
        collectionNum = tt[1:7]




    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")

    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("".join(collectionNum))
    # datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("化石、彩陶器、青铜礼器、彩绘陶器及百戏俑、彩绘乐舞俑、唐三彩等")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData77(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        museumPosition = re.findall(museumName, item)[3]
        museumPosition = list(museumPosition)
        del (museumPosition[0])
        lenname = len(museumPosition)
        del (museumPosition[lenname - 1])
        datalist.append("".join(museumPosition))
        # datalist.append("河南郑州")


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("12个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[8]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])

        # 藏品数量  collectionNum
        collectionNum = re.findall(museumName, item)[9]
        collectionNum = list(collectionNum)
        del (collectionNum[0])
        lenname = len(collectionNum)
        del (collectionNum[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    # datalist.append("现场免费领取参观券，讲解收费为50元/场")
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    datalist.append("".join(collectionNum))
    # datalist.append("Null")
    # 藏品种类  collectionType
    datalist.append("生产劳动画像、建筑类画像、历史故事类、社会生活类画像、天文与神话类等")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist

def getData78(baseurl,i):
    datalist = []

    url = baseurl
    html = askURL(url)
    # 2. 逐一解析数据
    soup = BeautifulSoup(html, "html.parser")
    # 查找符合要求的字符串
    # for item in soup.find_all('dd', class_="lemmaWgt-lemmaTitle-title"):  # class_属性值
    for item in soup.find_all('div', class_="basic-info cmn-clearfix"):
        item = str(item)
        # re 库用来通过正则表达式查找指定的字符
        # 博物馆编号
        # data.append(str(i))
        i = list(str(i))
        while len(i) < 3:
            i.insert(0, '0')
        datalist.append("".join(i))

        # 博物馆名称
        name = re.findall(museumName, item)[0]
        name = list(name)
        del (name[0])
        lenname = len(name)
        del (name[lenname - 1])
        datalist.append("".join(name))

        # 博物馆地点  museumPosition
        # museumPosition = re.findall(museumName, item)[3]
        # museumPosition = list(museumPosition)
        # del (museumPosition[0])
        # lenname = len(museumPosition)
        # del (museumPosition[lenname - 1])
        # datalist.append("".join(museumPosition))
        datalist.append("河南省开封市郑开大道 ")


        # 博物馆类别  buildingType
        buildingType = re.findall(museumName, item)[2]
        buildingType = list(buildingType)
        del (buildingType[0])
        lenname = len(buildingType)
        del (buildingType[lenname - 1])
        datalist.append("".join(buildingType))

        # 展区数量
        datalist.append("4个")

        # 建馆时间  buildingTime
        buildingTime = re.findall(museumName, item)[4]
        buildingTime = list(buildingTime)
        del (buildingTime[0])
        lenname = len(buildingTime)
        del (buildingTime[lenname - 1])
        datalist.append("".join(buildingTime))

        # 门票  ticket
        ticket = re.findall(museumName, item)[7]
        ticket = list(ticket)
        del (ticket[0])
        lenname = len(ticket)
        del (ticket[lenname - 1])


        # 开放时间  openingHours
        openingHours = re.findall(museumName, item)[5]
        openingHours = list(openingHours)
        del (openingHours[0])
        lenname = len(openingHours)
        del (openingHours[lenname - 1])

        # 藏品数量  collectionNum
        # collectionNum = re.findall(museumName, item)[7]
        # collectionNum = list(collectionNum)
        # del (collectionNum[0])
        # lenname = len(collectionNum)
        # del (collectionNum[lenname - 1])



    for item in soup.find_all('div', class_="summary-pic"):
        item = str(item)
        # 图片
        pic = re.findall(picture, item)[0]
        datalist.append(pic)

        # 宣传视频或音频  VideoAudio
        datalist.append("Null")


    # 门票  ticket
    datalist.append("".join(ticket))
    # 开放时间  openingHours
    datalist.append("".join(openingHours))


    # 藏品数量  collectionNum
    # datalist.append("近8万件")
    datalist.append("80000件")
    # 藏品种类  collectionType
    datalist.append("玉器、书画、青铜、陶瓷等")
    # 评价  evaluate
    datalist.append("环境优美，馆藏丰富。")
    # 评分  grade
    datalist.append("5分")

    # print("  datalist = ",datalist)
    return datalist


# 得到指定一个URL的网页内容
def askURL(url):
    # print("askURL")
    head = {
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.18362"
    }
                      # 用户代理：表示告诉豆瓣服务器，我们是什么类型的机器，浏览器（本质上是告诉浏览器，我们可以接受什么水平的文件
    request = urllib.request.Request(url,headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8")
        #print(html)
    except urllib.error.URLError as e:
        if hasattr(e,"code"):
            print(e.code)
        if hasattr(e,"reason"):
            print(e.reason)

    return html
    # 也能是列表
    #head[""]






# 3. 保存数据
import xlwt
def saveData(datalist,savepath,book,sheet,i):
    # print("cd")
    print("第 %d 条"%i)
    data = datalist
    for j in range(len(data)):
        sheet.write(i+1,j,data[j])  # 数据
    book.save(savepath)




    # sheet = book.add_sheet('豆瓣电影Top250',cell_overwrite_ok=True)
    # col = ("电影详情链接","图片链接","影片中文名","影片外文名","评价数")
    # for i in range(0,5):
    #     sheet.write(0,i,col[i]) # 列名
    # for i in range(0,250):
    #     print("第 %d 条"%i)
    #     data = datalist[i]
    #     for j in range(0, 5):
    #         sheet.write(i+1,j,data[j])  # 数据
    # book.save(savepath)



if __name__ == "__main__":
    main()
    #init_db("movietext.db")

    print("爬取完毕")




































